#!/usr/bin/env python3
import subprocess, sys, select
from datetime import datetime, timedelta
import json, os, time
from openpyxl import load_workbook

# --------------------- VALITSE JSON ---------------------
JSON_DIR = "./json_data"
JSON_FILE = os.path.join(JSON_DIR, "aircraft.json")
# Testisoftan käyttö (kommentoi yllä oleva ja ota tämä käyttöön)
# JSON_FILE = os.path.join(JSON_DIR, "aircraft_backup.json")

EXCEL_FILE = "./aircraftDatabase.xlsx"
REMOVE_TIMEOUT = 180
HEARTBEAT_INTERVAL = 5

planes_dict = {}
aircraft_db = {}
start_time = time.time()

# ---------------- LUE EXCEL (ICAO METADATA) ----------------
def load_aircraft_database():
    global aircraft_db
    if not os.path.exists(EXCEL_FILE):
        print("Excel metadata file not found.")
        return
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active
    headers = [str(cell.value).strip() for cell in sheet[1]]
    header_index = {name.lower(): headers.index(name) for name in headers}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_icao = row[header_index["icao24"]]
        if raw_icao is None:
            continue
        icao = str(raw_icao).strip().lower()
        aircraft_db[icao] = {
            "registration": str(row[header_index.get("registration", 0)] or "??").strip(),
            "typecode": str(row[header_index.get("typecode", 0)] or "??").strip(),
            "operator": str(row[header_index.get("operator", 0)] or "??").strip(),
        }

# ---------------- HELPER ----------------
def get_state(p):
    return (
        round(p.get("altitude", 0), 0),
        round(p.get("speed", 0), 0),
        round(p.get("track", 0), 0),
        round(p.get("lat", 5), 5),
        round(p.get("lon", 5), 5),
        round(p.get("rssi", 0.0), 1)
    )

def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")

def format_runtime(seconds):
    td = timedelta(seconds=int(seconds))
    return str(td)

# ---------------- LADATAAN EXCEL ----------------
load_aircraft_database()

# ---------------- KÄYNNISTETÄÄN DUMP1090 ----------------
print("Käynnistetään dump1090...")
proc = subprocess.Popen(
    ["dump1090-mutability", "--net", "--write-json", JSON_DIR],
    stdout=subprocess.PIPE,
    stderr=subprocess.STDOUT,
    text=True
)

# ---------------- LUE TIKUN “JORINAT” 6 SEKUNTIA ----------------
print("\nLuen tikun alkuraportteja 6 sekuntia...")
initial_start = time.time()
while time.time() - initial_start < 6:
    if proc.stdout in select.select([proc.stdout], [], [], 0.1)[0]:
        line = proc.stdout.readline()
        if line:
            print(line.strip())

# ---------------- DASHBOARD ALKU ----------------
clear_screen()
print("="*160)
print(f"FLIGHT TRANSPONDER – RIKASTETTU NÄKYMÄ    {datetime.now()}")
print(f"Running time: 0:00:00")
print("="*160)
print(f"{'ICAO':<8} {'Reg':<10} {'Type':<6} {'Operator':<20} {'Flight':<8} "
      f"{'Lat':>8} {'Lon':>8} {'Alt':>6} {'Spd':>5} {'Track':>6} {'RSSI':>6} "
      f"{'Status':>9} {'Removed':>9}")
print("-"*160)
print("Odottamassa koneita JSONista...")
print("="*160)
time.sleep(1)

# ---------------- PÄÄSILMUKKA ----------------
last_heartbeat = time.time()
try:
    while True:
        now = time.time()
        seen_this_round = set()
        runtime_str = format_runtime(now - start_time)

        # Lue JSON
        if os.path.exists(JSON_FILE):
            try:
                with open(JSON_FILE,"r") as f:
                    data = json.load(f)
                planes = data.get("aircraft", [])
            except json.JSONDecodeError:
                planes = []
        else:
            planes = []

        # Päivitä koneet
        for p in planes:
            raw_icao = p.get("hex","?")
            try:
                icao = f"{int(raw_icao):06X}".lower()  # numerosta heksa, esim. 780092 -> aa3487
            except ValueError:
                icao = str(raw_icao).strip().lower()

            flight = p.get("flight","?")
            state = get_state(p)
            seen_this_round.add(icao)

            if icao not in planes_dict:
                meta = aircraft_db.get(icao, {"registration":"??","typecode":"??","operator":"??"})
                planes_dict[icao] = {
                    "flight": flight,
                    "state": state,
                    "first_seen": datetime.now().strftime("%H:%M:%S"),
                    "last_seen": now,
                    "status": "ACTIVE",
                    "removed_time": None,
                    "registration": meta["registration"],
                    "type": meta["typecode"],
                    "operator": meta["operator"]
                }
            else:
                planes_dict[icao]["flight"] = flight
                planes_dict[icao]["state"] = state
                planes_dict[icao]["last_seen"] = now
                if planes_dict[icao]["status"] == "REMOVED":
                    planes_dict[icao]["status"] = "ACTIVE"
                    planes_dict[icao]["removed_time"] = None

        # Tarkista poistuneet
        for icao, pdata in planes_dict.items():
            if pdata["status"] == "ACTIVE" and icao not in seen_this_round:
                if now - pdata["last_seen"] > REMOVE_TIMEOUT:
                    pdata["status"] = "REMOVED"
                    pdata["removed_time"] = datetime.now().strftime("%H:%M:%S")

        # Dashboard tulostus
        clear_screen()
        print("="*160)
        print(f"FLIGHT TRANSPONDER – RIKASTETTU NÄKYMÄ    {datetime.now()}")
        print(f"Running time: {runtime_str}")
        print("="*160)
        print(f"{'ICAO':<8} {'Reg':<10} {'Type':<6} {'Operator':<20} {'Flight':<8} "
              f"{'Lat':>8} {'Lon':>8} {'Alt':>6} {'Spd':>5} {'Track':>6} {'RSSI':>6} "
              f"{'Status':>9} {'Removed':>9}")
        print("-"*160)

        if not planes_dict:
            if now - last_heartbeat >= HEARTBEAT_INTERVAL:
                print("Ei havaintoja JSONista – ohjelma toimii.")
                last_heartbeat = now
        else:
            for icao, pdata in planes_dict.items():
                s = pdata["state"]
                removed_time = pdata["removed_time"] or "-"
                print(f"{icao:<8} {pdata['registration']:<10} {pdata['type']:<6} {pdata['operator']:<20} "
                      f"{pdata['flight']:<8} {s[3]:>8} {s[4]:>8} {s[0]:>6} {s[1]:>5} {s[2]:>6} {s[5]:>6} "
                      f"{pdata['status']:>9} {removed_time:>9}")

            print("="*160)
            print(f"Yhteensä päivän aikana nähtyjä koneita: {len(planes_dict)}")

        time.sleep(1)

except KeyboardInterrupt:
    clear_screen()
    print("\nLopetetaan ohjelma...")
    proc.terminate()
    proc.wait()